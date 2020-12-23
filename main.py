# 30 pourcent a preciser sur fichier retour
# ajout verif case vide dans partie fr
# faire fct exte modif fichier
# faire fichier config
# code 14 editer fichier texte
# preciser deans fichier erreur case marron reduc
# centrer references et origines es et fr

import os
import csv
import pandas as pd
import numpy as np
import sys
import fct_tk
import time

from tkinter import Tk    
from tkinter.filedialog import askopenfilename
from os import path as os_path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill

dossier_python = os_path.abspath(os_path.split(__file__)[0])

##############################################################
##Emplacement colonnes

ref_fr = 0   #### A=0, B=1, etc...
nom_trad = 1
ori_trad = 2
nom_fr = 3
ori_fr = 4
prix_fr = 6
reduc1_fr = 8
reduc2_fr = 9
reduc3_fr = 10

ref_es = 0
prix_es = 3
nom_es = 2
ori_es = 8####################
reduc1_es = 4
reduc2_es = 5
reduc3_es = 6


##############################################################
##Fonctions

def ecriture(liste):

    for element in liste :

        myfile.write(str(element))
        myfile.write('\n')


def isDigit(x):

    try:

        float(x)
        return True

    except ValueError:

        return False


def comp_nom (n_es,n_fr,mot,ref,erreur,x):

    if (str(n_es)).strip() != (str(n_fr)).strip() :

        erreur.append(' '+ref+'\n  FR = '+n_fr+"\n  ES = "+n_es+'\n')

        if erreur == erreur8 :
            #print ('ok')
            #print (x)
            #print (nom_fr)
            df_fr.loc[x,str(nom_fr)]= n_es
            a = chr(ord('@')+nom_fr+1)
            b = x+2
            c = chr(ord('@')+nom_trad+1)
            liste_vert.append(a+str(b))
            liste_jaune.append(c+str(b))

        elif erreur == erreur13 :

            df_fr.loc[x,str(ori_fr)]= n_es
            a = chr(ord('@')+ori_fr+1)
            b = x+2
            c = chr(ord('@')+ori_trad+1)
            liste_vert.append(a+str(b))
            liste_jaune.append(c+str(b))


def comp_prix (p_es,p_fr, ref, n_fr, n_es):

    a = chr(ord('@')+prix_fr+1)
    b = x+2

    if p_es == 0 :

        erreur9.append('  '+ref+' '+n_fr)
        liste_marron.append(a+str(b))

    elif p_es != 0 and p_es != p_fr :

        if (str(n_es)).strip() == (str(n_fr)).strip() :

             erreur10.append('\n '+ref+' '+n_fr[:40]+"\n  ancien prix  "+str(p_fr)+"\n  nouveau prix "+str(p_es))

        else :

            erreur10.append('\n '+ref+'\n  nom FR : '+n_fr+'\n  nom ES : '+n_es+'\n  prix FR '+str(p_fr)+'\n  prix ES '+str(p_es))

        if p_es/p_fr >= 1.3 or p_es/p_fr <= 0.7 :

            liste_marron.append(a+str(b))

        else :

            df_fr.loc[x,str(prix_fr)]= p_es
            liste_vert.append(a+str(b))
     


def comp_promo (prom_es,prom_fr, n_fr, ref, num):

    b = x+2

    if prom_es != 0 and not (prom_fr == 'X' or prom_fr == 'x') :

        erreur11.append('  '+ref+' '+n_fr+' promo '+num)

        if num == '1' :
            
            a = chr(ord('@')+reduc1_fr+1)
            df_fr.loc[x,str(reduc1_fr)]= 'X'
            liste_vert.append(a+str(b))
        
        if num == '2' :
            
            a = chr(ord('@')+reduc2_fr+1)
            df_fr.loc[x,str(reduc2_fr)]= 'X'
            liste_vert.append(a+str(b))

        if num == '3' :
            
            a = chr(ord('@')+reduc3_fr+1)
            df_fr.loc[x,str(reduc3_fr)]= 'X'
            liste_vert.append(a+str(b))


    elif prom_es == 0 and (str(prom_fr).strip() == 'X' or str(prom_fr).strip() == 'x'):

        a = chr(ord('@')+reduc1_fr+1)
        erreur12.append('  '+ref+' '+n_fr+' promo '+num)
        liste_marron.append(a+str(b))


def rech_nom (dfObj, value, coll): 

    listOfPos = [] 
    result = dfObj.isin([value]) 
    rows = list(result[coll][result[coll] == True].index) 
  
    for row in rows: 

        listOfPos.append((row)) 

    if (len(listOfPos)) == 0 :

        return 'empty'

    elif (len(listOfPos)) > 1 :

        return 'doublon' 

    else :

        return listOfPos[0]


def comp_class ():

    comp_prix (dic_es['Prix_es'],dic_fr['Prix_fr'], dic_fr['R_fr'], dic_fr['N_fr'],dic_es['N_es'])
    comp_nom (dic_es['N_es'],dic_fr['N_fr'],'nom', dic_fr['R_fr'], erreur8,x)
    comp_nom (dic_es['Ori_es'],dic_fr['Ori_fr'],'origine',dic_fr['R_fr'], erreur13,x)
    comp_promo (dic_es['Reduc_es1'],dic_fr['Reduc_fr1'], dic_fr['N_fr'],dic_fr['R_fr'],'1')
    comp_promo (dic_es['Reduc_es2'],dic_fr['Reduc_fr2'], dic_fr['N_fr'],dic_fr['R_fr'],'2')
    comp_promo (dic_es['Reduc_es3'],dic_fr['Reduc_fr3'], dic_fr['N_fr'],dic_fr['R_fr'],'3')


##############################################################
## Choix fichier es 

input("\nAppuyer sur Entree pour choisir le fichier du catalogue espagnol\n")

os.chdir(dossier_python)

Tk().withdraw() 
File_es = askopenfilename()

read_file_es = fct_tk.choix_fichier(File_es)
#File = 'es.xlsx'#########3
#
#read_file = pd.read_excel(File,sheet_name=0, skiprows = 1, header=None) ##########

read_file_es.to_csv ("es.csv", 
                  index = None, 
                  header = True)

df_es = pd.DataFrame(pd.read_csv("es.csv")) 


##############################################################
## Choix fichier fr

input("\nAppuyer sur Entree pour choisir le fichier du catalogue francais non modifie\n")
 
Tk().withdraw() 
File_fr = askopenfilename()

read_file_fr = fct_tk.choix_fichier(File_fr)

#File = 'sample_fr.xlsx'
#
#read_file = pd.read_excel(File,sheet_name='BDD', skiprows = 1, header=None) 

os.chdir(dossier_python)

read_file_fr.to_csv ("fr.csv", 
                  index = None, 
                  header = True)

df_fr = pd.DataFrame(pd.read_csv("fr.csv")) 


###########################################
## Nettoyage fichier esp

refa_es = df_es.columns.values[ref_es] 
prixa_es = df_es.columns.values[prix_es] 
noma_es = df_es.columns.values[nom_es] 
reduca1_es = df_es.columns.values[reduc1_es]
reduca2_es = df_es.columns.values[reduc2_es]
reduca3_es = df_es.columns.values[reduc3_es]

check = pd.isna(df_es[refa_es])
check2 = pd.isna(df_es[prixa_es])
check3 = pd.isna(df_es[noma_es])

for x in range(len(df_es.index)):
    if check[x] == True :   # si ref vide : ref=vide
        df_es.iat[x,ref_es] = 'vide'

for x in range(len(df_es.index)):
    if check[x] == True and check2[x] == True : # si prix et ref vide suppr ligne
        df_es = df_es.drop([x])
    
for x in range(len(df_es.index)):# si produit vide suppr ligne
    if check3[x] == True :
        df_es = df_es.drop([x])

df_es[prixa_es] = pd.to_numeric(df_es[prixa_es],errors='coerce') # si prix = str => 0
df_es = df_es.replace(np.nan, 0, regex=True)
df_es[prixa_es] = df_es[prixa_es].astype(float)

df_es[reduca1_es] = pd.to_numeric(df_es[reduca1_es],errors='coerce') # si prix = str => 0
df_es = df_es.replace(np.nan, 0, regex=True)
df_es[reduca1_es] = df_es[reduca1_es].astype(float)

df_es[reduca2_es] = pd.to_numeric(df_es[reduca2_es],errors='coerce') # si prix = str => 0
df_es = df_es.replace(np.nan, 0, regex=True)
df_es[reduca2_es] = df_es[reduca2_es].astype(float)

df_es[reduca3_es] = pd.to_numeric(df_es[reduca3_es],errors='coerce') # si prix = str => 0
df_es = df_es.replace(np.nan, 0, regex=True)
df_es[reduca3_es] = df_es[reduca3_es].astype(float)

df_es = df_es.drop_duplicates(subset=[prixa_es, refa_es, noma_es])# supr des vrais doublons
################################# VERIF PROMO
df_es[prixa_es] = round((df_es[prixa_es] * 1.37),2)

list2 = [] ### reconstruction de l'index

for x in range (len(df_es.index)):
    list2.append(x)

df_es.index = list2


#####################################
#### recherche duplicata fichier esp

list_es = df_es[refa_es].tolist()

dupl_es = [] 

if not len(list_es) == len(set(list_es)):
    for i in list_es:
        if list_es.count(i) > 1 and i != 'vide':
            dupl_es.append(i)



##############################################################
## Verif fichier fr

refa_fr = df_fr.columns.values[ref_fr]
prixa_fr = df_fr.columns.values[prix_fr] 
noma_fr = df_fr.columns.values[nom_fr]

df_fr[prixa_fr] = pd.to_numeric(df_fr[prixa_fr])  ## convertion des prix, arret programme erreur 
df_fr[prixa_fr] = df_fr[prixa_fr].astype(float)
df_fr[prixa_fr] = round((df_fr[prixa_fr]),2)

if (df_fr[refa_fr].isnull) == True :
    print ("\nReference manquante dans le catalogue fr")
    sys.exit()

list_fr = df_fr[refa_fr].tolist()
dupl_fr = set() 

if not len(list_fr) == len(set(list_fr)):
    # ~ print ("\nDoublon de reference cote fr ")
    for i in list_fr:
        if list_fr.count(i) > 1:
            dupl_fr.add(i)

##############################################################
##  Comparaison

liste_vert = []
liste_marron = []
liste_jaune = []

liste_rouge = []#ligne
liste_gris = []#ligne
liste_noire = []#ligne
liste_violet = []#ligne

liste_erreur1 = []

erreur1 = []
erreur2 = []
erreur3 = []
erreur4 = []
erreur5 = []
erreur6 = []
erreur7 = []
erreur8 = []
erreur9 = []
erreur10 = []
erreur11 = []
erreur12 = []
erreur13 = []
erreur14 = []

def fdico_es () :
    dico_es  =  {
                'N_es' : df_es.iat[a,nom_es],
                'R_es' : df_es.iat[a,ref_es],
                'Prix_es' : df_es.iat[a,prix_es],
                'Ori_es' : df_es.iat[a,ori_es],
                'Reduc_es1' : df_es.iat[a,reduc1_es],
                'Reduc_es2' : df_es.iat[a,reduc2_es],
                'Reduc_es3' : df_es.iat[a,reduc3_es]
                }
    return dico_es

def fdico_fr () :
    dico_fr = {
                'N_fr' : df_fr.iat[x,nom_fr],
                'R_fr' : df_fr.iat[x,ref_fr],
                'Prix_fr' : df_fr.iat[x,prix_fr],
                'Ori_fr' : df_fr.iat[x,ori_fr],
                'Reduc_fr1' : df_fr.iat[x,reduc1_fr],
                'Reduc_fr2' : df_fr.iat[x,reduc2_fr],
                'Reduc_fr3' : df_fr.iat[x,reduc3_fr]
                }
    return dico_fr


#########################################################################
######################## Boucle

for x in dupl_fr :  ### Verif si les doublons fr sont aussi des doublons es

    if not x in dupl_es :

        erreur1.append("Produit ref "+x)
        liste_erreur1.append(x)

for x in range (len(df_fr.index)) :

    dic_fr = fdico_fr() ### Redef des valeurs avec changement x


    if dic_fr['R_fr'] in liste_erreur1 :

        liste_rouge.append(x+1)


    elif dic_fr['R_fr'] in dupl_es : ## Recherche si ref est en doublon cote esp

        a = rech_nom (df_es, dic_fr['N_fr'], noma_es) ## Recherche par nom

        if a == 'empty' :

            erreur2.append('  '+dic_fr['R_fr']+' '+dic_fr['N_fr'])
            liste_rouge.append(x+1)

        elif a == 'doublon' :

            erreur3.append('  '+dic_fr['R_fr']+' '+dic_fr['N_fr'])
            liste_rouge.append(x+1)

        else :

            dic_es = fdico_es()

            if dic_fr['R_fr'] == dic_es['R_es']:

                comp_class ()              #### Comparaison classique

            else :

                erreur14.append(' REF FR = '+dic_fr['R_fr']+'\n REF ES = '+dic_es['R_es']+'\n '+dic_fr['N_fr'])
                liste_rouge.append(x+1)

    else :

        a = rech_nom (df_es, dic_fr['R_fr'], refa_es) ## Recherche ref cote esp
        
        if a == 'empty' :      ## Echec recherche

            a = rech_nom (df_es, dic_fr['N_fr'], noma_es)    ## Recherche nom

            if a == 'empty' :        ## Echec

                erreur4.append('  '+dic_fr['R_fr']+' '+dic_fr['N_fr'])
                liste_gris.append(x+1)

            elif a == 'doublon' :

                erreur5.append('  '+dic_fr['R_fr']+' '+dic_fr['N_fr'])
                liste_rouge.append(x+1)

            else :  

                dic_es = fdico_es()                             
                erreur6.append('  Ref FR = '+dic_fr['R_fr']+' ' +dic_fr['N_fr'][:20]+'\n'+'  Ref ES = '+dic_es['R_es'])

                a = chr(ord('@')+ref_fr+1)
                df_fr.loc[x,str(ref_fr)]= dic_es['R_es']
                b = x+2
                liste_vert.append(a+str(b))

                comp_class()  #### Comparaison classique
                
        elif a == 'doublon' :

            erreur7.append("!!!!!!!!erreur anormale dans la detection des doublons (a signaler), la comparaison pour le produit"+dic_fr['R_fr'])
            liste_noire.append(x+1) 

        else : 

            dic_es = fdico_es()  
            comp_class ()####possible erreur prix   #### Comparaison classique


from fct_erreur import error

error(erreur1,erreur2,erreur3,erreur4,erreur5,erreur6,erreur7,erreur8,erreur9,erreur10,erreur11,erreur12,erreur13,erreur14)


##################################################
### Coloriage fichier Excel


def coloriage(liste,couleur):

    for element in liste:

        d = ws[element]
        d.fill = couleur
        d.font = police

def coloriage_ligne(liste, couleur):

    for element in liste:

        for cell in list(ws.rows)[element]:

             cell.fill = couleur
             cell.font = police

wb = load_workbook(filename = File_fr)

wb.security.workbookPassword = '1ZERTY7'

ws = wb['BDD']

ws.delete_rows(2,ws.max_row)


jaune = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type='solid')

rouge = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type='solid')

vert = PatternFill(
       start_color="00FF00",
       end_color="00FF00",
       fill_type='solid')

gris = PatternFill(
       start_color="808080",
       end_color="808080",
       fill_type='solid')

noire = PatternFill(
       start_color="000000",
       end_color="000000",
       fill_type='solid')

marron = PatternFill(
       start_color="800000",
       end_color="800000",
       fill_type='solid')

police = Font(
        name = "Arial",
        size = 10)


for r in dataframe_to_rows(df_fr, index=None, header=None):
    ws.append(r)

coloriage(liste_vert, vert)
coloriage (liste_jaune, jaune)
coloriage (liste_marron, marron)

coloriage_ligne(liste_gris, gris)
coloriage_ligne(liste_noire, noire)
coloriage_ligne(liste_rouge, rouge)

wb.save(filename = 'retour.xlsx')


### Font en arial 8 et refaire alignement ref

##############################################################################
#################

os.remove("fr.csv")

print("\nComparaison terminee, resultats sauves dans retour.txt et retour.xlsx dans le dossier contenant le script\n")

time.sleep(1)


#######################################
#######################################

rep = int(input('Taper 1 pour creer un tableur comportant tous les produits du catalogue espagnol non presents dans le catalogue francais sinon taper 2\n'))

if rep != 1:
    os.remove("es.csv")
    sys.exit()

df3 = df_es

for a in range (len(df_es.index)):
    dic_es = fdico_es()
    x= rech_nom (df_fr, dic_es['N_es'], noma_fr)
    if x != 'empty':
        df3 = df3.drop([a])
    if df_es.iat[a,prix_es] == 0 : 
        df3 = df3.drop([a])

df_es = df3

df_es.drop(df_es.columns[[1,7,9,10,11,12]], axis = 1, inplace = True) 
df_es.columns =['Ref', 'Nom', 'Prix', 'Promo','Promo2','Promo3','Origine'] 
df_es = df_es.sort_values('Nom')

df_es['Promo'] = round((df_es['Promo']),2)
df_es['Promo2'] = round((df_es['Promo2']),2)
df_es['Promo3'] = round((df_es['Promo3']),2)

df_es.to_excel('Comparaison.xlsx', engine='openpyxl', index = False)  
os.remove("es.csv")

print("\nComparaison terminee, resultat sauve dans Comparaison.xlsx dans le dosier contenant le script\n")

input("Appuyer sur entree pour quitter le programme")




